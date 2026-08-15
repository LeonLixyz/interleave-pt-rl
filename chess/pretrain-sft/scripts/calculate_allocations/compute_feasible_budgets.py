import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
BASE_DIR = "compute_allocation_tables/C_total_02_2p994e18"
INPUT_CSV = "02_allocation_data_table.csv"
INPUT_CSV = os.path.join(BASE_DIR, INPUT_CSV)
OUTPUT_XLSX = "allocation_summary_for_doc.xlsx"
OUTPUT_XLSX = os.path.join(BASE_DIR, OUTPUT_XLSX)

PRETRAIN_SHARD_CAP = 5860
RL_STEP_CAP = 1000


def human_format_tokens(x: float) -> str:
    if pd.isna(x):
        return ""
    x = float(x)
    abs_x = abs(x)

    if abs_x >= 1e12:
        return f"{x / 1e12:.3f}T"
    if abs_x >= 1e9:
        return f"{x / 1e9:.3f}B"
    if abs_x >= 1e6:
        return f"{x / 1e6:.3f}M"
    if abs_x >= 1e3:
        return f"{x / 1e3:.3f}K"
    return f"{x:.3f}"


def round3(x):
    if pd.isna(x):
        return x
    return round(float(x), 3)


def yes_no(x) -> str:
    return "Yes" if bool(x) else "No"


def build_summary_table(df: pd.DataFrame) -> pd.DataFrame:
    required_cols = [
        "model",
        "beta",
        "alpha",
        "gamma",
        "sft_total_tokens",
        "sft_total_samples_all_epochs",
        "pretrain_tokens_actual",
        "pretrain_shards",
        "pretrain_steps_approx",
        "rl_total_processed_tokens",
        "rl_estimated_gradient_steps",
        "rl_num_prompts",
        "sft_feasible_under_80k_per_epoch",
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    out = pd.DataFrame()
    out["model size"] = df["model"]
    out["beta"] = df["beta"].map(round3)
    out["alpha"] = df["alpha"].map(round3)
    out["gamma"] = df["gamma"].map(round3)

    out["sft total token"] = df["sft_total_tokens"].map(human_format_tokens)
    out["sft number samples"] = df["sft_total_samples_all_epochs"].map(round3)

    out["pretrain total token"] = df["pretrain_tokens_actual"].map(human_format_tokens)
    out["pretrain shard"] = df["pretrain_shards"].astype(int)
    out["pretrain step"] = df["pretrain_steps_approx"].astype(int)

    out["rl total processed tokens"] = df["rl_total_processed_tokens"].map(human_format_tokens)
    out["rl estimated gradient steps"] = df["rl_estimated_gradient_steps"].map(round3)
    out["rl num prompts"] = df["rl_num_prompts"].map(round3)

    out["sft feasibility"] = df["sft_feasible_under_80k_per_epoch"].map(yes_no)
    out["pretrain feasibility"] = (df["pretrain_shards"] <= PRETRAIN_SHARD_CAP).map(yes_no)
    out["rl feasibility"] = (df["rl_estimated_gradient_steps"] <= RL_STEP_CAP).map(yes_no)

    model_order = {"50M": 50, "100M": 100, "200M": 200, "300M": 300, "500M": 500}
    out["_model_order"] = out["model size"].map(model_order).fillna(10**9)

    out = out.sort_values(
        by=["_model_order", "beta", "alpha", "gamma"],
        ascending=[True, True, True, True],
    ).drop(columns="_model_order").reset_index(drop=True)

    return out


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 12,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 18,
        "F": 18,
        "G": 20,
        "H": 14,
        "I": 14,
        "J": 20,
        "K": 24,
        "L": 16,
        "M": 14,
        "N": 18,
        "O": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    green_fill = PatternFill("solid", fgColor="E2F0D9")
    red_fill = PatternFill("solid", fgColor="FDE9E7")

    headers = [cell.value for cell in ws[1]]
    feasibility_cols = []
    for idx, h in enumerate(headers, start=1):
        if h in {"sft feasibility", "pretrain feasibility", "rl feasibility"}:
            feasibility_cols.append(idx)

    for row in range(2, ws.max_row + 1):
        for col in feasibility_cols:
            cell = ws.cell(row=row, column=col)
            if cell.value == "Yes":
                cell.fill = green_fill
            elif cell.value == "No":
                cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")


def main(input_csv=INPUT_CSV, output_xlsx=OUTPUT_XLSX):
    df = pd.read_csv(input_csv)
    summary_df = build_summary_table(df)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="summary")

    wb = load_workbook(output_xlsx)
    ws = wb["summary"]
    style_worksheet(ws)
    wb.save(output_xlsx)

    print(f"Saved to {output_xlsx}")


if __name__ == "__main__":
    main()